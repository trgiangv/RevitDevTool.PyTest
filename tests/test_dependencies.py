"""Dependency tests resolved from suite metadata in conftest.py."""

import pytest
from functools import wraps
import humanize
import networkx as nx
import numpy as np
import polars as pl
from openpyxl import Workbook
from pydantic import BaseModel, ValidationError
from shapely.geometry import Polygon, box
from tabulate import tabulate


requires_document = pytest.mark.skipif(
    "__revit__" not in dir() or __revit__.ActiveUIDocument is None,  # noqa: F821
    reason="No document open in Revit",
)


def test_humanize_from_conftest(humanize_mod):
    """humanize was declared in conftest.py and auto-installed."""
    result = humanize_mod.naturalsize(1_048_576)
    assert "1.0 MB" in result
    print(f"humanize.naturalsize(1MB) = {result}")

    result_time = humanize_mod.naturaldelta(3661)
    assert "hour" in result_time
    print(f"humanize.naturaldelta(3661s) = {result_time}")


def test_humanize_intcomma():
    """humanize.intcomma formats large numbers."""
    assert humanize.intcomma(1_000_000) == "1,000,000"
    assert humanize.intcomma(1234567890) == "1,234,567,890"
    print(f"humanize {humanize.__version__} — intcomma OK")


def test_tabulate_basic():
    """tabulate is available from suite dependency setup."""
    data = [["Wall", 42], ["Floor", 15], ["Roof", 3]]
    table = tabulate(data, headers=["Element", "Count"], tablefmt="grid")
    assert "Wall" in table
    assert "42" in table
    print(f"tabulate output:\n{table}")


@requires_document
def test_tabulate_with_revit_data():
    """Combine tabulate (auto-installed) with Revit API data."""
    from Autodesk.Revit.DB import FilteredElementCollector, BuiltInCategory

    doc = __revit__.ActiveUIDocument.Document  # noqa: F821

    categories = [
        ("Walls", BuiltInCategory.OST_Walls),
        ("Floors", BuiltInCategory.OST_Floors),
        ("Roofs", BuiltInCategory.OST_Roofs),
        ("Doors", BuiltInCategory.OST_Doors),
        ("Windows", BuiltInCategory.OST_Windows),
    ]
    rows = []
    for name, cat in categories:
        count = FilteredElementCollector(doc).OfCategory(cat).WhereElementIsNotElementType().GetElementCount()
        rows.append([name, count])

    table = tabulate(rows, headers=["Category", "Count"], tablefmt="simple")
    assert len(rows) == 5
    print(f"Element summary:\n{table}")


def test_numpy_linalg():
    """numpy: matrix operations from Pixi environment."""
    a = np.array([[1, 2], [3, 4]], dtype=np.float64)
    b = np.linalg.inv(a)
    identity = a @ b
    assert np.allclose(identity, np.eye(2))
    print(f"numpy {np.__version__} — matrix inverse OK")


def test_polars_groupby():
    """polars: DataFrame from Pixi environment."""
    df = pl.DataFrame({
        "category": ["A", "B", "A", "B", "A"],
        "value": [10, 20, 30, 40, 50],
    })
    result = df.group_by("category").agg(pl.col("value").sum()).sort("category")
    sums = result["value"].to_list()
    assert sums == [90, 60]
    print(f"polars {pl.__version__} — group_by OK")


def test_shapely_intersection():
    """shapely: geometry operations (auto-installed via conftest.py PEP 723)."""
    poly = Polygon([(0, 0), (10, 0), (10, 10), (0, 10)])
    clip = box(5, 5, 15, 15)
    intersection = poly.intersection(clip)
    assert intersection.area == pytest.approx(25.0)
    print(f"shapely — intersection area={intersection.area}")


def test_pydantic_validation():
    """pydantic: model validation from Pixi environment."""
    class WallInfo(BaseModel):
        name: str
        length: float
        height: float

    wall = WallInfo(name="W-01", length=10.5, height=3.0)
    assert wall.model_dump() == {"name": "W-01", "length": 10.5, "height": 3.0}

    with pytest.raises(ValidationError):
        WallInfo(name="bad", length="not_a_number", height=3.0)

    print("pydantic — validation + serialization OK")


def test_networkx_shortest_path():
    """networkx: graph algorithms (auto-installed via conftest.py PEP 723)."""
    G = nx.Graph()
    G.add_weighted_edges_from([(1, 2, 1), (2, 3, 2), (1, 3, 10), (3, 4, 1)])

    path = nx.shortest_path(G, source=1, target=4, weight="weight")
    assert path == [1, 2, 3, 4]
    print(f"networkx {nx.__version__} — shortest path OK")


def test_openpyxl_workbook():
    """openpyxl: in-memory Excel workbook (auto-installed via conftest.py PEP 723)."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Element"
    ws["B1"] = "Count"
    ws["A2"] = "Walls"
    ws["B2"] = 42
    assert ws["B2"].value == 42
    print("openpyxl — workbook OK")
